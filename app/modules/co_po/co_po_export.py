"""Export helpers for CO-PO Attainment reports: PDF (WeasyPrint) and
multi-sheet XLSX (openpyxl, one sheet per table) — both deferred-import so
a missing native dependency can't break app boot."""


def render_pdf(setup: dict, result: dict) -> bytes:
    from flask import render_template
    from weasyprint import HTML
    html_string = render_template("co_po/pdf.html", setup=setup, result=result)
    return HTML(string=html_string).write_pdf()


def render_xlsx(setup: dict, result: dict) -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")

    def style_header(ws):
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

    # ── Sheet 1: CO Attainment ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "CO Attainment"
    ws1.append(["CO", "Description", "Target", "Attained", "Status"])
    style_header(ws1)
    for co in result["co_attainment"]:
        ws1.append([
            co["co_id"], co["description"], co["target"],
            co["attainment"] if co["attainment"] is not None else "No data",
            "Met" if co["met"] else ("Below Target" if co["attainment"] is not None else "No Data"),
        ])
    for col, width in zip("ABCDE", [8, 50, 10, 12, 16]):
        ws1.column_dimensions[col].width = width

    # ── Sheet 2: CO-PO Mapping Matrix ────────────────────────────────
    ws2 = wb.create_sheet("CO-PO Matrix")
    po_ids = [po["po_id"] for po in result["po_attainment"]]
    ws2.append(["CO"] + po_ids)
    style_header(ws2)
    po_mapping = setup.get("po_mapping", {})
    for co in setup.get("course_outcomes", []):
        row = [co["id"]] + [po_mapping.get(co["id"], {}).get(po_id, "") for po_id in po_ids]
        ws2.append(row)
    ws2.column_dimensions["A"].width = 8

    # ── Sheet 3: PO Attainment ───────────────────────────────────────
    ws3 = wb.create_sheet("PO Attainment")
    ws3.append(["PO", "Name", "Attainment (0-3)", "Level"])
    style_header(ws3)
    for po in result["po_attainment"]:
        ws3.append([
            po["po_id"], po["po_name"],
            po["attainment"] if po["attainment"] is not None else "No data",
            "Below Threshold" if po["level"] == "red" else "OK",
        ])
    for col, width in zip("ABCD", [8, 36, 16, 16]):
        ws3.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
