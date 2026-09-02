"""
Teacher Analytics (Module 13) + Topic Difficulty Heatmap (Module 17).

Reads across collections written by other modules (attendance, quiz,
tests, dropout, wellness) without writing to any of them, except the
heatmap's "Generate Re-teach Material" action which reuses the Study
Material pipeline. All aggregation happens in Python rather than Mongo
aggregation pipelines — the class sizes this app targets (tens of
students) make that both simpler to test against mongomock and fast
enough in practice.
"""
from collections import defaultdict
from datetime import datetime, timedelta

from bson import ObjectId
from bson.errors import InvalidId

from app.extensions import db, logger
from app.modules.dropout.prediction_service import RECOMMENDATIONS

SCORE_BUCKETS = ["0-20", "21-40", "41-60", "61-80", "81-100"]
GRADE_ORDER = ["A+", "A", "B", "C", "F"]
DATE_RANGE_DAYS = {"7d": 7, "30d": 30, "semester": 120, "all": None}
LOW_ATTENDANCE_THRESHOLD = 75


# ═══════════════════════════════════════════════════════════════════
#  Shared lookups
# ═══════════════════════════════════════════════════════════════════

def _users_by_id() -> dict:
    return {str(u["_id"]): u for u in db.users.find({"role": "student"})}


def _students_by_id() -> dict:
    return {s["student_id"]: s for s in db.students.find()}


def _tests_by_id() -> dict:
    out = {}
    for t in db.tests.find():
        out[str(t["_id"])] = t
    return out


def _bucket_for_score(pct: float) -> str:
    if pct < 40:
        return "red"
    if pct < 65:
        return "amber"
    return "green"


def _date_cutoff(range_key: str):
    days = DATE_RANGE_DAYS.get(range_key, None)
    if not days:
        return None
    return datetime.utcnow() - timedelta(days=days)


# ═══════════════════════════════════════════════════════════════════
#  Attendance Analytics
# ═══════════════════════════════════════════════════════════════════

def get_attendance_analytics() -> dict:
    students = list(db.students.find())
    logs = list(db.attendance_logs.find())

    sessions_held = len({log["timestamp"].date() for log in logs if log.get("timestamp")})

    bar_chart = {
        "labels": [s.get("name", s.get("student_id")) for s in students][:30],
        "chart_values": [s.get("total_attendance", 0) for s in students][:30],
    }

    by_day = defaultdict(int)
    for log in logs:
        ts = log.get("timestamp")
        if ts:
            by_day[ts.date().isoformat()] += 1
    trend_days = sorted(by_day.keys())[-30:]
    trend_chart = {"labels": trend_days, "chart_values": [by_day[d] for d in trend_days]}

    low_attendance = []
    for s in students:
        pct = round(min(100, (s.get("total_attendance", 0) / sessions_held) * 100), 1) if sessions_held else 0
        if pct < LOW_ATTENDANCE_THRESHOLD:
            low_attendance.append({
                "student_id": s["student_id"], "name": s.get("name", s["student_id"]),
                "total_attendance": s.get("total_attendance", 0), "percentage": pct,
            })
    low_attendance.sort(key=lambda r: r["percentage"])

    return {
        "sessions_held": sessions_held,
        "bar_chart": bar_chart,
        "trend_chart": trend_chart,
        "low_attendance": low_attendance,
        "students": students,
    }


# ═══════════════════════════════════════════════════════════════════
#  Quiz Analytics
# ═══════════════════════════════════════════════════════════════════

def get_quiz_analytics() -> dict:
    results = list(db.quiz_results.find())
    users = _users_by_id()

    score_buckets = {b: 0 for b in SCORE_BUCKETS}
    grade_counts = {g: 0 for g in GRADE_ORDER}
    by_topic = defaultdict(list)
    by_user = defaultdict(list)

    for r in results:
        pct = r.get("score_percent", 0)
        if pct <= 20:
            score_buckets["0-20"] += 1
        elif pct <= 40:
            score_buckets["21-40"] += 1
        elif pct <= 60:
            score_buckets["41-60"] += 1
        elif pct <= 80:
            score_buckets["61-80"] += 1
        else:
            score_buckets["81-100"] += 1

        grade = r.get("grade", "F")
        if grade in grade_counts:
            grade_counts[grade] += 1

        by_topic[r.get("topic", "Unknown")].append(pct)
        by_user[r.get("user_id")].append(pct)

    topic_table = [
        {"topic": topic, "average": round(sum(scores) / len(scores), 1), "attempts": len(scores)}
        for topic, scores in by_topic.items()
    ]
    topic_table.sort(key=lambda t: t["average"])

    ranking = []
    for user_id, scores in by_user.items():
        user = users.get(user_id)
        ranking.append({
            "name": user.get("name", "Unknown") if user else "Unknown",
            "average": round(sum(scores) / len(scores), 1),
            "attempts": len(scores),
        })
    ranking.sort(key=lambda r: r["average"], reverse=True)

    return {
        "total_quizzes": len(results),
        "score_buckets": score_buckets,
        "grade_counts": grade_counts,
        "topic_table": topic_table,
        "ranking": ranking[:20],
    }


# ═══════════════════════════════════════════════════════════════════
#  Dropout Analytics
# ═══════════════════════════════════════════════════════════════════

def get_dropout_analytics() -> dict:
    all_predictions = list(db.dropout_predictions.find().sort("created_at", -1))
    users = _users_by_id()

    latest_by_user = {}
    for p in all_predictions:
        uid = p.get("user_id")
        if uid and uid not in latest_by_user:
            latest_by_user[uid] = p

    risk_counts = {"Low": 0, "Medium": 0, "High": 0}
    high_risk_list = []
    for uid, p in latest_by_user.items():
        level = p.get("risk_level", "Low")
        if level in risk_counts:
            risk_counts[level] += 1
        if level == "High":
            user = users.get(uid)
            high_risk_list.append({
                "name": user.get("name", "Unknown") if user else "Unknown",
                "probability": p.get("probability", 0),
                "recommendations": RECOMMENDATIONS.get("High", [])[:3],
            })
    high_risk_list.sort(key=lambda r: r["probability"], reverse=True)

    return {
        "risk_counts": risk_counts,
        "high_risk_list": high_risk_list,
        "total_assessed": len(latest_by_user),
    }


# ═══════════════════════════════════════════════════════════════════
#  Wellness Analytics (anonymised — no student names anywhere here)
# ═══════════════════════════════════════════════════════════════════

def get_wellness_analytics() -> dict:
    assessments = list(db.wellness_assessments.find().sort("timestamp", -1))
    latest_by_user = {}
    for a in assessments:
        uid = a.get("user_id")
        if uid and uid not in latest_by_user:
            latest_by_user[uid] = a

    severity_counts = {"Minimal": 0, "Mild": 0, "Moderate": 0, "Severe": 0}
    for a in latest_by_user.values():
        level = a.get("severity", "Minimal")
        if level in severity_counts:
            severity_counts[level] += 1

    sessions = list(db.wellness_sessions.find())
    by_day = defaultdict(int)
    crisis_dates = []
    for s in sessions:
        created = s.get("created_at")
        if created:
            by_day[created.date().isoformat()] += 1
        if s.get("crisis_flagged") and created:
            crisis_dates.append(created.date().isoformat())
    trend_days = sorted(by_day.keys())[-30:]

    return {
        "severity_counts": severity_counts,
        "sessions_trend": {"labels": trend_days, "chart_values": [by_day[d] for d in trend_days]},
        "total_sessions": len(sessions),
        "crisis_count": len(crisis_dates),
        "crisis_dates": sorted(set(crisis_dates), reverse=True)[:10],
    }


# ═══════════════════════════════════════════════════════════════════
#  Topic Difficulty Heatmap (Module 17)
# ═══════════════════════════════════════════════════════════════════

def _collect_topic_records(date_from=None, subject=None, branch=None, year=None) -> list[dict]:
    users = _users_by_id()
    students = _students_by_id()
    tests = _tests_by_id()
    records = []

    query = {}
    if date_from:
        query["timestamp"] = {"$gte": date_from}
    for r in db.quiz_results.find(query):
        user = users.get(r.get("user_id"))
        u_branch = user.get("branch") if user else None
        u_year = user.get("year") if user else None
        if subject and r.get("subject") != subject:
            continue
        if branch and u_branch != branch:
            continue
        if year and str(u_year) != str(year):
            continue
        records.append({
            "topic": r.get("topic", "Unknown"), "subject": r.get("subject", "General"),
            "score_percent": r.get("score_percent", 0), "timestamp": r.get("timestamp"),
            "identity": r.get("user_id"), "identity_type": "user",
        })

    attempt_query = {"grading_status": "complete"}
    if date_from:
        attempt_query["submitted_at"] = {"$gte": date_from}
    for a in db.test_attempts.find(attempt_query):
        test = tests.get(a.get("test_id"))
        if not test:
            continue
        student = students.get(a.get("student_id"))
        s_branch = student.get("branch") if student else None
        s_year = student.get("year") if student else None
        if subject and test.get("subject") != subject:
            continue
        if branch and s_branch != branch:
            continue
        if year and str(s_year) != str(year):
            continue
        total_marks = a.get("total_marks") or 0
        pct = round((a.get("score", 0) / total_marks) * 100, 1) if total_marks else 0
        records.append({
            "topic": test.get("title", "Untitled Test"), "subject": test.get("subject", "General"),
            "score_percent": pct, "timestamp": a.get("submitted_at"),
            "identity": a.get("student_id"), "identity_type": "student",
        })

    return records


def get_available_subjects() -> list[str]:
    subjects = set(db.quiz_results.distinct("subject"))
    subjects |= {t.get("subject") for t in db.tests.find({}, {"subject": 1}) if t.get("subject")}
    return sorted(s for s in subjects if s)


def get_heatmap_data(date_range="all", subject=None, branch=None, year=None) -> dict:
    date_from = _date_cutoff(date_range)
    records = _collect_topic_records(date_from, subject, branch, year)

    by_topic = defaultdict(list)
    for r in records:
        by_topic[(r["subject"], r["topic"])].append(r)

    tiles = []
    for (subj, topic), recs in by_topic.items():
        scores = [r["score_percent"] for r in recs]
        dates = [r["timestamp"] for r in recs if r["timestamp"]]
        avg = round(sum(scores) / len(scores), 1)
        tiles.append({
            "subject": subj, "topic": topic, "average": avg, "attempts": len(recs),
            "bucket": _bucket_for_score(avg),
            "date_from": min(dates).strftime("%b %d, %Y").replace(" 0", " ") if dates else "—",
            "date_to": max(dates).strftime("%b %d, %Y").replace(" 0", " ") if dates else "—",
        })
    tiles.sort(key=lambda t: (t["subject"], t["topic"]))

    red_flags = sorted([t for t in tiles if t["average"] < 50], key=lambda t: t["average"])

    by_week = defaultdict(list)
    for r in records:
        if r["timestamp"]:
            iso_year, iso_week, _ = r["timestamp"].isocalendar()
            by_week[f"{iso_year}-W{iso_week:02d}"].append(r["score_percent"])
    weeks = sorted(by_week.keys())[-10:]
    trend_chart = {
        "labels": weeks,
        "chart_values": [round(sum(by_week[w]) / len(by_week[w]), 1) for w in weeks],
    }

    return {
        "tiles": tiles, "red_flags": red_flags, "trend_chart": trend_chart,
        "filters": {"date_range": date_range, "subject": subject, "branch": branch, "year": year},
        "record_count": len(records),
    }


def get_heatmap_drilldown(subject: str, topic: str, date_range="all", branch=None, year=None) -> dict:
    date_from = _date_cutoff(date_range)
    records = _collect_topic_records(date_from, subject, branch, year)
    records = [r for r in records if r["topic"] == topic]

    users = _users_by_id()
    students = _students_by_id()

    by_identity = defaultdict(list)
    for r in records:
        by_identity[(r["identity_type"], r["identity"])].append(r)

    rows = []
    attempted_ids = set()
    for (id_type, identity), recs in by_identity.items():
        if id_type == "user":
            person = users.get(identity)
            name = person.get("name", "Unknown") if person else "Unknown"
            attempted_ids.add(("user", identity))
        else:
            person = students.get(identity)
            name = person.get("name", identity) if person else identity
            attempted_ids.add(("student", identity))
        scores = [r["score_percent"] for r in recs]
        dates = [r["timestamp"] for r in recs if r["timestamp"]]
        rows.append({
            "name": name, "attempts": len(recs),
            "average": round(sum(scores) / len(scores), 1),
            "best": max(scores) if scores else 0,
            "last_attempt": max(dates).strftime("%b %d, %Y").replace(" 0", " ") if dates else "—",
        })
    rows.sort(key=lambda r: r["average"])

    zero_attempt = [
        s.get("name", s["student_id"]) for s in students.values()
        if ("student", s["student_id"]) not in attempted_ids
    ]

    return {"topic": topic, "subject": subject, "rows": rows, "zero_attempt": sorted(zero_attempt)}


def trigger_reteach(teacher_id: str, subject: str, topic: str, branch: str, year: str) -> dict:
    from app.modules.nlp import study_service
    return study_service.run_pipeline(teacher_id, topic, subject, branch or "ECE", year or "3")


def render_heatmap_pdf(heatmap: dict) -> bytes:
    from flask import render_template
    from weasyprint import HTML
    html_string = render_template("analytics/heatmap_pdf.html", heatmap=heatmap)
    return HTML(string=html_string).write_pdf()


# ═══════════════════════════════════════════════════════════════════
#  Exports
# ═══════════════════════════════════════════════════════════════════

def export_attendance_xlsx() -> bytes:
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    data = get_attendance_analytics()
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1D4ED8", end_color="1D4ED8", fill_type="solid")

    headers = ["Student ID", "Name", "Branch", "Year", "Total Attendance", "Attendance %"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    for s in data["students"]:
        pct = round(min(100, (s.get("total_attendance", 0) / data["sessions_held"]) * 100), 1) if data["sessions_held"] else 0
        ws.append([
            s.get("student_id", ""), s.get("name", ""), s.get("branch", ""),
            s.get("year", ""), s.get("total_attendance", 0), pct,
        ])

    for col, width in zip("ABCDEF", [14, 24, 10, 8, 16, 14]):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_quiz_csv() -> str:
    import csv
    import io
    data = get_quiz_analytics()
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Rank", "Student", "Average Score %", "Quizzes Taken"])
    for i, r in enumerate(data["ranking"], start=1):
        writer.writerow([i, r["name"], r["average"], r["attempts"]])
    writer.writerow([])
    writer.writerow(["Topic", "Average Score %", "Attempts"])
    for t in data["topic_table"]:
        writer.writerow([t["topic"], t["average"], t["attempts"]])
    return buf.getvalue()
